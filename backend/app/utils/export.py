import csv
import io
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side


def export_csv(headers: List[str], rows: List[List[Any]]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue()


def export_xlsx(headers: List[str], rows: List[List[Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
