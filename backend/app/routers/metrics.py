"""
指标管理系统 API
"""
import os
import io
from fastapi import APIRouter, Depends, Query, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from urllib.parse import quote
from sqlalchemy import select, or_, func, case
from sqlalchemy.ext.asyncio import AsyncSession

from app.deps.db_deps import get_db
from app.deps.auth_deps import require_process, require_developer, get_current_user
from app.core.response import success, paginated, error
from app.utils.pagination import paginate
from app.schemas.metrics import (
    IndicatorCreateReq, IndicatorUpdateReq, IndicatorResp,
    IndicatorBatchUpdateReq,
    BomIndicatorCreateReq, BomIndicatorUpdateReq, BomIndicatorResp,
    BomIndicatorBatchCreateReq, BomIndicatorBatchUpdateReq, BomIndicatorBatchStatusReq,
    BomIndicatorParamAddReq, BomIndicatorParamUpdateReq,
    BomConfigCreateReq, BomConfigUpdateReq, BomConfigResp,
    BomDomainOwnersReq, BomDomainOwnersResp,
    BomConfigCopyReq,
    CollectionCreateReq, CollectionUpdateReq, CollectionResp,
    CollectionTestItemCreateReq, CollectionTestItemUpdateReq, CollectionTestItemResp, CollectionTestItemOwnerUpdateReq,
    IndicatorParamAddReq, IndicatorParamUpdateReq,
    BomIndicatorParamAddReq, BomIndicatorParamUpdateReq,
    TestItemIndicatorCreateReq, TestItemIndicatorBatchCreateReq, TestItemIndicatorResp,
    ScriptTemplateCreateReq, ScriptTemplateUpdateReq, ScriptTemplateResp,
    ScriptExecuteReq, ScriptExecuteResp,
    BomExportReq,
    ReviewReq, ReviewActionResp,
    BomIndicatorBatchSaveReq, BomIndicatorBatchSaveResp,
    ParamChangeLogResp,
    OnlineUserInfo,
    RollbackReq,
    VersionSnapshotResp,
)
from app.services.dictionary_service import DictionaryService
from app.services.collection_service import CollectionService
from app.services.bom_config_service import BomConfigService
from app.services.version_snapshot_service import VersionSnapshotService
from app.services.script_template_service import ScriptTemplateService
from app.models.metrics import (
    BomConfig, BomIndicator, IndicatorDict,
    TestItemCollection, CollectionTestItem, IndicatorVersionSnapshot, ScriptTemplate,
)
from app.utils.export import export_xlsx

router = APIRouter(tags=["指标管理"])

dict_svc = DictionaryService()
coll_svc = CollectionService()
bom_svc = BomConfigService()
ver_svc = VersionSnapshotService()


def _get_operator(user: dict) -> str:
    return user.get("display_name") or user.get("username", "")


# ════════════════════════════════════════════
# 指标字典库
# ════════════════════════════════════════════
@router.get("/indicators", dependencies=[Depends(require_process)])
async def list_indicators(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    keyword: str = Query(""),
    category: str = Query(""),
    status: int = Query(None),
    db: AsyncSession = Depends(get_db),
):
    items, total, p, ps = await dict_svc.list(
        db, page=page, page_size=page_size, keyword=keyword, category=category, status=status,
    )
    return paginated([IndicatorResp(**i.to_dict()) for i in items], total, p, ps)


@router.get("/indicators/all", dependencies=[Depends(require_process)])
async def list_all_indicators(db: AsyncSession = Depends(get_db)):
    items = await dict_svc.list_all_active(db)
    return success(data=[IndicatorResp(**i.to_dict()) for i in items])


@router.get("/indicators/categories", dependencies=[Depends(require_process)])
async def list_indicator_categories(db: AsyncSession = Depends(get_db)):
    cats = await dict_svc.list_categories(db)
    return success(data=cats)


@router.get("/indicators/domains", dependencies=[Depends(require_process)])
async def list_indicator_domains(db: AsyncSession = Depends(get_db)):
    data = await dict_svc.list_domains(db)
    return success(data=data)


@router.post("/indicators", dependencies=[Depends(require_developer)])
async def create_indicator(req: IndicatorCreateReq, db: AsyncSession = Depends(get_db)):
    obj = await dict_svc.create(db, req.model_dump())
    return success(data=IndicatorResp(**obj.to_dict()), message="指标创建成功")


@router.put("/indicators/{indicator_id}", dependencies=[Depends(require_developer)])
async def update_indicator(
    indicator_id: int, req: IndicatorUpdateReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await dict_svc.update(db, indicator_id, req.model_dump(exclude_none=True), operator=_get_operator(user))
    return success(data=IndicatorResp(**obj.to_dict()), message="指标更新成功")


@router.delete("/indicators/{indicator_id}", dependencies=[Depends(require_developer)])
async def delete_indicator(indicator_id: int, force: bool = Query(False), db: AsyncSession = Depends(get_db)):
    await dict_svc.delete(db, indicator_id, force=force)
    return success(message="指标已停用")


@router.put("/indicators/batch", dependencies=[Depends(require_developer)])
async def batch_update_indicators(req: IndicatorBatchUpdateReq, db: AsyncSession = Depends(get_db)):
    objs = await dict_svc.batch_update(db, [item.model_dump(exclude_none=True) for item in req.items])
    return success(data=[IndicatorResp(**o.to_dict()) for o in objs], message=f"已更新 {len(objs)} 条指标")


# ── Per-indicator Script ──


@router.get("/indicators/{indicator_id}/script", dependencies=[Depends(require_process)])
async def get_indicator_script(indicator_id: int, db: AsyncSession = Depends(get_db)):
    source = await dict_svc.get_script(db, indicator_id)
    default_script = dict_svc.get_default_script()
    has_custom = source != default_script
    return success(data={"source_code": source, "has_custom": has_custom, "is_default": not has_custom})


@router.put("/indicators/{indicator_id}/script", dependencies=[Depends(require_developer)])
async def update_indicator_script(
    indicator_id: int,
    data: dict = Body(...),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    source_code = data.get("source_code", "")
    if not source_code.strip():
        return error(code=400, message="脚本代码不能为空")
    try:
        await dict_svc.update_script(db, indicator_id, source_code, operator=_get_operator(user))
        return success(data={"source_code": source_code}, message="脚本保存成功")
    except ValueError as e:
        return error(code=400, message=str(e))


@router.post("/indicators/{indicator_id}/script/validate", dependencies=[Depends(require_process)])
async def validate_indicator_script(data: dict = Body(...)):
    source_code = data.get("source_code", "")
    result = DictionaryService.validate_script(source_code)
    return success(data=result)


@router.post("/indicators/{indicator_id}/script/preview", dependencies=[Depends(require_process)])
async def preview_indicator_script(
    indicator_id: int,
    data: dict = Body(...),
    db: AsyncSession = Depends(get_db),
):
    source_code = data.get("source_code", "")
    input_data = data.get("input_data", {})
    if not source_code.strip():
        return error(code=400, message="脚本代码为空")
    result = await dict_svc.preview_script(db, indicator_id, source_code, input_data)
    return success(data=result)


@router.post("/indicators/{indicator_id}/script/reset", dependencies=[Depends(require_developer)])
async def reset_indicator_script(
    indicator_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    await dict_svc.reset_script(db, indicator_id, operator=_get_operator(user))
    return success(data={"source_code": dict_svc.get_default_script()}, message="脚本已重置为默认")


@router.post("/indicators/import", dependencies=[Depends(require_developer)])
async def import_indicators_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        return error(code=400, message="仅支持 .xlsx / .xls 文件")
    content = await file.read()
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    if not ws:
        return error(code=400, message="Excel 文件为空")
    headers_map = {"指标编码": "code", "指标名称": "name", "分类": "category", "单位": "unit", "描述": "description"}
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header_row = [c.value for c in list(ws.iter_rows(min_row=1, max_row=1))[0]] if ws.max_column else []
    col_map = {header_row[i]: i for i in range(len(header_row)) if header_row[i] in headers_map}

    imported = 0
    errors = []
    for r_idx, row in enumerate(rows, 2):
        try:
            data = {}
            for cn, field in headers_map.items():
                col_idx = col_map.get(cn)
                if col_idx is None:
                    continue
                val = row[col_idx] if col_idx < len(row) else None
                data[field] = str(val or "").strip()
            if not data.get("code") or not data.get("name"):
                errors.append(f"第 {r_idx} 行: 指标编码或名称为空")
                continue
            existing = await db.execute(select(IndicatorDict).where(IndicatorDict.code == data["code"]))
            if existing.scalar_one_or_none():
                errors.append(f"第 {r_idx} 行: 指标编码 '{data['code']}' 已存在")
                continue
            await dict_svc.create(db, data)
            imported += 1
        except Exception as e:
            errors.append(f"第 {r_idx} 行: {str(e)}")
    msg = f"成功导入 {imported} 条"
    if errors:
        msg += f"，{len(errors)} 条跳过:\n" + "\n".join(errors[:10])
        if len(errors) > 10:
            msg += f"\n...及另外 {len(errors) - 10} 条"
    return success(message=msg)


@router.put("/indicators/save", dependencies=[Depends(require_developer)])
async def save_indicator(
    data: dict,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    indicator_id = data.get("id")
    if indicator_id:
        obj = await dict_svc.update(db, indicator_id, data, operator=_get_operator(user))
        return success(data=IndicatorResp(**obj.to_dict()), message="指标更新成功")
    else:
        obj = await dict_svc.create(db, data)
        return success(data=IndicatorResp(**obj.to_dict()), message="指标创建成功")


@router.get("/indicators/alerts", dependencies=[Depends(require_process)])
async def list_alert_indicators(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    keyword: str = Query(""),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(IndicatorDict)
    if keyword:
        stmt = stmt.where(
            IndicatorDict.name.ilike(f"%{keyword}%") | IndicatorDict.code.ilike(f"%{keyword}%")
        )
    stmt = stmt.order_by(IndicatorDict.id.desc())
    items, total, p, ps = await paginate(db, stmt, page, page_size)

    # Enrich with reference counts
    result = []
    for it in items:
        d = it.to_dict()
        ti_r = await db.execute(
            select(func.count()).select_from(TestItemIndicator)
            .where(TestItemIndicator.indicator_id == it.id)
        )
        d["ref_test_items"] = ti_r.scalar() or 0
        bi_r = await db.execute(
            select(func.count()).select_from(BomIndicator)
            .where(BomIndicator.indicator_id == it.id)
        )
        d["ref_bom_configs"] = bi_r.scalar() or 0
        result.append(d)
    return paginated(result, total, p, ps)


@router.get("/indicators/{indicator_id}/references", dependencies=[Depends(require_process)])
async def get_indicator_references(indicator_id: int, db: AsyncSession = Depends(get_db)):
    refs = await dict_svc.get_references(db, indicator_id)
    return success(data={
        "indicator_id": indicator_id,
        "collections": refs["collections"],
        "bom_configs": refs["bom_configs"],
        "total_collections": len(refs["collections"]),
        "total_bom_configs": len(refs["bom_configs"]),
    })


# ── Dictionary-level Per-param CRUD (test_params) ──


@router.post("/indicators/{indicator_id}/params", dependencies=[Depends(require_developer)])
async def add_indicator_param(indicator_id: int, req: IndicatorParamAddReq, db: AsyncSession = Depends(get_db)):
    try:
        obj = await dict_svc.add_param(db, indicator_id, req.model_dump())
    except ValueError as e:
        return error(code=400, message=str(e))
    return success(data=IndicatorResp(**obj.to_dict()).model_dump())


@router.put("/indicators/{indicator_id}/params/{param_key}", dependencies=[Depends(require_developer)])
async def update_indicator_param(indicator_id: int, param_key: str, req: IndicatorParamUpdateReq, db: AsyncSession = Depends(get_db)):
    data = {k: v for k, v in req.model_dump().items() if v is not None}
    obj = await dict_svc.update_param(db, indicator_id, param_key, data)
    return success(data=IndicatorResp(**obj.to_dict()).model_dump())


@router.delete("/indicators/{indicator_id}/params/{param_key}", dependencies=[Depends(require_developer)])
async def delete_indicator_param(indicator_id: int, param_key: str, db: AsyncSession = Depends(get_db)):
    obj = await dict_svc.delete_param(db, indicator_id, param_key)
    return success(data=IndicatorResp(**obj.to_dict()).model_dump())


# ════════════════════════════════════════════
# 测试项集合管理
# ════════════════════════════════════════════
@router.get("/collections", dependencies=[Depends(require_process)])
async def list_collections(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    keyword: str = Query(""),
    status: int = Query(None),
    db: AsyncSession = Depends(get_db),
):
    items, total, p, ps = await coll_svc.list(db, page=page, page_size=page_size, keyword=keyword, status=status)
    return paginated([CollectionResp(**i.to_dict()) for i in items], total, p, ps)


@router.get("/collections/all", dependencies=[Depends(require_process)])
async def list_all_collections(db: AsyncSession = Depends(get_db)):
    items = await coll_svc.list_all_active(db)
    return success(data=[CollectionResp(**i.to_dict()) for i in items])


@router.get("/collections/{collection_id}/available-indicators", dependencies=[Depends(require_process)])
async def list_collection_available_indicators(collection_id: int, db: AsyncSession = Depends(get_db)):
    items = await coll_svc.list_collection_available_indicators(db, collection_id)
    return success(data=items)


@router.post("/collections", dependencies=[Depends(require_developer)])
async def create_collection(req: CollectionCreateReq, db: AsyncSession = Depends(get_db)):
    obj = await coll_svc.create(db, req.model_dump())
    return success(data=CollectionResp(**obj.to_dict()), message="集合创建成功")


@router.put("/collections/{collection_id}", dependencies=[Depends(require_developer)])
async def update_collection(
    collection_id: int, req: CollectionUpdateReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await coll_svc.update(db, collection_id, req.model_dump(exclude_none=True), operator=_get_operator(user))
    return success(data=CollectionResp(**obj.to_dict()), message="集合更新成功")


@router.delete("/collections/{collection_id}", dependencies=[Depends(require_developer)])
async def archive_collection(
    collection_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    await coll_svc.update_status(db, collection_id, 0, operator=_get_operator(user))
    from app.models.metrics import TestItemCollection
    from sqlalchemy import select
    r = await db.execute(select(TestItemCollection).where(TestItemCollection.id == collection_id))
    coll = r.scalar_one_or_none()
    if coll:
        coll.version = (coll.version or 0) + 1
        await db.flush()
    return success(message="集合已归档")


@router.get("/collections/{collection_id}/items", dependencies=[Depends(require_process)])
async def list_collection_items(
    collection_id: int,
    version: int = Query(None, description="集合版本号，传此参数时从快照中获取对应版本的测试项"),
    db: AsyncSession = Depends(get_db),
):
    if version is not None:
        from app.models.metrics import IndicatorVersionSnapshot
        r = await db.execute(
            select(IndicatorVersionSnapshot)
            .where(
                IndicatorVersionSnapshot.entity_type == "collection",
                IndicatorVersionSnapshot.entity_id == collection_id,
                IndicatorVersionSnapshot.version == version,
            )
            .order_by(IndicatorVersionSnapshot.created_at.desc())
            .limit(1)
        )
        snapshot = r.scalar_one_or_none()
        if snapshot:
            items = snapshot.snapshot_data.get("items", [])
            # 协同编辑字段（乐观锁版本号/负责人）属于运行时状态，需用实时数据覆盖快照中的旧值
            from app.models.metrics import CollectionTestItem as _CTI
            ids = [i.get("id") for i in items if i.get("id")]
            live_map = {}
            if ids:
                lr = await db.execute(select(_CTI).where(_CTI.id.in_(ids)))
                for ti in lr.scalars().all():
                    live_map[ti.id] = ti
            for i in items:
                ti = live_map.get(i.get("id"))
                if ti:
                    i["item_revision"] = ti.item_revision
                    i["owner_id"] = ti.owner_id
                    i["owner_name"] = ti.owner_name
                    i["owner_manual"] = bool(ti.owner_manual)
            # 领域由实时指标绑定推导，快照项也统一补齐
            domains = await coll_svc.get_item_domains(db, ids)
            for i in items:
                i.setdefault("domain", domains.get(i.get("id"), ""))
            return success(data=[CollectionTestItemResp(**i) for i in items])
    items = await coll_svc.list_items_with_domain(db, collection_id)
    return success(data=[CollectionTestItemResp(**i) for i in items])


@router.post("/collections/{collection_id}/items", dependencies=[Depends(require_developer)])
async def create_collection_item(
    collection_id: int, req: CollectionTestItemCreateReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await coll_svc.create_item(db, collection_id, req.model_dump(), operator=_get_operator(user))
    return success(data=CollectionTestItemResp(**obj.to_dict()), message="测试项添加成功")


@router.put("/collections/items/{item_id}", dependencies=[Depends(require_developer)])
async def update_collection_item(
    item_id: int, req: CollectionTestItemUpdateReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await coll_svc.update_item(db, item_id, req.model_dump(exclude_none=True), operator=_get_operator(user))
    return success(data=CollectionTestItemResp(**obj.to_dict()), message="测试项更新成功")


@router.put("/collections/items/{item_id}/owner", dependencies=[Depends(require_developer)])
async def update_collection_item_owner(
    item_id: int,
    req: CollectionTestItemOwnerUpdateReq = Body(...),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """单行手动修改测试项负责人（局部覆盖领域自动填充规则）"""
    obj = await bom_svc.update_item_owner(db, item_id, owner_name=req.owner_name, operator=_get_operator(user))
    return success(data=CollectionTestItemResp(**obj.to_dict()), message="测试项负责人已更新")


@router.delete("/collections/items/{item_id}", dependencies=[Depends(require_developer)])
async def delete_collection_item(
    item_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    await coll_svc.delete_item(db, item_id, operator=_get_operator(user))
    return success(message="测试项已删除")


# ── Test Item Indicators (二期: 测试项绑定指标) ──
@router.get("/collections/items/{item_id}/indicators", dependencies=[Depends(require_process)])
async def list_item_indicators(item_id: int, db: AsyncSession = Depends(get_db)):
    items = await coll_svc.list_item_indicators(db, item_id)
    return success(data=[TestItemIndicatorResp(**i) for i in items])


@router.post("/collections/items/{item_id}/indicators/batch", dependencies=[Depends(require_developer)])
async def batch_add_item_indicators(
    item_id: int, req: TestItemIndicatorBatchCreateReq,
    db: AsyncSession = Depends(get_db),
):
    await coll_svc.batch_add_item_indicators(db, item_id, [r.model_dump() for r in req.indicators])
    items = await coll_svc.list_item_indicators(db, item_id)
    return success(data=[TestItemIndicatorResp(**i) for i in items], message="指标绑定成功")


@router.delete("/collections/items/indicators/{indicator_id}", dependencies=[Depends(require_developer)])
async def delete_item_indicator(indicator_id: int, db: AsyncSession = Depends(get_db)):
    await coll_svc.delete_item_indicator(db, indicator_id)
    return success(message="指标已解绑")


# ════════════════════════════════════════════
# BOM 指标配置
# ════════════════════════════════════════════
@router.get("/bom-configs", dependencies=[Depends(require_process)])
async def list_bom_configs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    keyword: str = Query(""),
    status: int = Query(None),
    db: AsyncSession = Depends(get_db),
):
    items, total, p, ps = await bom_svc.list(db, page=page, page_size=page_size, keyword=keyword, status=status)
    data = []
    for i in items:
        d = i.to_dict()
        d["is_latest"] = True
        if i.collection_id:
            # Compare with max snapshot version for this collection (BOM only references, never writes)
            r = await db.execute(
                select(func.max(IndicatorVersionSnapshot.version))
                .where(
                    IndicatorVersionSnapshot.entity_type == "collection",
                    IndicatorVersionSnapshot.entity_id == i.collection_id,
                )
            )
            max_ver = r.scalar()
            if max_ver is not None:
                d["is_latest"] = i.collection_version >= max_ver
            d["collection_version"] = i.collection_version
        data.append(d)
    return paginated(data, total, p, ps)


@router.get("/bom-configs/bom-codes", dependencies=[Depends(require_process)])
async def list_bom_codes(keyword: str = Query(""), db: AsyncSession = Depends(get_db)):
    codes = await bom_svc.list_bom_codes(db, keyword)
    return success(data=codes)


async def _enrich_bom_configs(db: AsyncSession, items):
    data = []
    for i in items:
        d = i.to_dict()
        d["is_latest"] = True
        if i.collection_id:
            r = await db.execute(
                select(func.max(IndicatorVersionSnapshot.version))
                .where(
                    IndicatorVersionSnapshot.entity_type == "collection",
                    IndicatorVersionSnapshot.entity_id == i.collection_id,
                )
            )
            max_ver = r.scalar()
            if max_ver is not None:
                d["is_latest"] = i.collection_version >= max_ver
            d["collection_version"] = i.collection_version
        data.append(d)
    return data


@router.get("/bom-configs/grouped", dependencies=[Depends(require_process)])
async def list_grouped_bom_configs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    keyword: str = Query(""),
    status: int = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """按 BOM 编码聚合：同一编码仅返回最新版本基础信息，历史版本移入编码指标页。"""
    items, total, p, ps = await bom_svc.list_grouped_by_code(db, page=page, page_size=page_size, keyword=keyword, status=status)
    data = await _enrich_bom_configs(db, items)
    codes = [d["bom_code"] for d in data]
    if codes:
        r = await db.execute(
            select(BomConfig.bom_code, func.count(BomConfig.id))
            .where(BomConfig.bom_code.in_(codes))
            .group_by(BomConfig.bom_code)
        )
        counts = dict(r.all())
        for d in data:
            d["version_count"] = counts.get(d["bom_code"], 1)
    return paginated(data, total, p, ps)


@router.get("/bom-configs/by-code", dependencies=[Depends(require_process)])
async def list_bom_configs_by_code(bom_code: str = Query(""), db: AsyncSession = Depends(get_db)):
    items = await bom_svc.list_by_code(db, bom_code)
    data = await _enrich_bom_configs(db, items)
    return success(data=data)


@router.get("/bom-configs/domain-owners", dependencies=[Depends(require_process)])
async def get_bom_domain_owners_by_bom_code(bom_code: str = Query(...), db: AsyncSession = Depends(get_db)):
    """获取 BOM 编码级别的领域负责人配置（所有版本共享）"""
    data = await bom_svc.get_domain_owners_by_bom_code(db, bom_code)
    return success(data=BomDomainOwnersResp(**data))


@router.put("/bom-configs/domain-owners", dependencies=[Depends(require_developer)])
async def update_bom_domain_owners_by_bom_code(
    bom_code: str = Query(...),
    req: BomDomainOwnersReq = Body(...),
    db: AsyncSession = Depends(get_db),
):
    """更新 BOM 编码级别的领域负责人配置（所有版本共享）"""
    owners = await bom_svc.update_domain_owners_by_bom_code(db, bom_code, req.domain_owners)
    return success(data={"domain_owners": owners}, message="领域负责人配置已保存")


@router.get("/bom-configs/{config_id}", dependencies=[Depends(require_process)])
async def get_bom_config(config_id: int, db: AsyncSession = Depends(get_db)):
    obj = await bom_svc.get(db, config_id)
    d = obj.to_dict()
    d["collection_version"] = obj.collection_version
    return success(data=BomConfigResp(**d))


@router.post("/bom-configs", dependencies=[Depends(require_developer)])
async def create_bom_config(req: BomConfigCreateReq, db: AsyncSession = Depends(get_db)):
    obj = await bom_svc.create(db, req.model_dump())
    return success(data=BomConfigResp(**obj.to_dict()), message="BOM配置创建成功")


@router.put("/bom-configs/{config_id}", dependencies=[Depends(require_developer)])
async def update_bom_config(config_id: int, req: BomConfigUpdateReq, db: AsyncSession = Depends(get_db)):
    obj = await bom_svc.update(db, config_id, req.model_dump(exclude_none=True))
    return success(data=BomConfigResp(**obj.to_dict()), message="BOM配置更新成功")


@router.delete("/bom-configs/{config_id}", dependencies=[Depends(require_developer)])
async def delete_bom_config(config_id: int, db: AsyncSession = Depends(get_db)):
    await bom_svc.delete(db, config_id)
    return success(message="BOM配置已删除")


@router.get("/bom-configs/{config_id}/domain-owners", dependencies=[Depends(require_process)])
async def get_bom_domain_owners(config_id: int, db: AsyncSession = Depends(get_db)):
    data = await bom_svc.get_domain_owners(db, config_id)
    return success(data=BomDomainOwnersResp(**data))


@router.put("/bom-configs/{config_id}/domain-owners", dependencies=[Depends(require_developer)])
async def update_bom_domain_owners(
    config_id: int,
    req: BomDomainOwnersReq,
    db: AsyncSession = Depends(get_db),
):
    owners = await bom_svc.update_domain_owners(db, config_id, req.domain_owners)
    return success(data={"domain_owners": owners}, message="领域负责人配置已保存")


@router.put("/bom-configs/{config_id}/switch-version", dependencies=[Depends(require_developer)])
async def switch_bom_collection_version(
    config_id: int,
    snapshot_id: int = Query(..., description="目标版本快照ID"),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """Switch the collection version referenced by a BOM config.
    This ONLY updates the BOM's reference — it does NOT modify the collection itself.
    Published or archived BOM configs must not have their collection version changed.
    """
    config = await bom_svc.get(db, config_id)
    if config.archived or config.review_status == "approved":
        raise ValueError("已发布或归档的 BOM 不能切换集合版本")
    # Validate the snapshot exists for this collection
    r = await db.execute(
        select(IndicatorVersionSnapshot)
        .where(
            IndicatorVersionSnapshot.id == snapshot_id,
            IndicatorVersionSnapshot.entity_type == "collection",
            IndicatorVersionSnapshot.entity_id == config.collection_id,
        )
    )
    snapshot = r.scalar_one_or_none()
    if not snapshot:
        raise NotFoundError("版本快照不存在或不属于当前集合")
    # Only update the BOM's reference, never touch the collection
    config.collection_version = snapshot.version
    await db.flush()
    await db.refresh(config)
    return success(data=BomConfigResp(**config.to_dict()), message=f"已切换至版本 v{snapshot.version}")


@router.get("/bom-configs/check-version", dependencies=[Depends(require_process)])
async def check_bom_version(
    bom_code: str = Query(..., description="BOM编码"),
    exclude_config_id: int = Query(0, description="排除的配置ID（编辑时排除自身）"),
    db: AsyncSession = Depends(get_db),
):
    """Check if a bom_code already has a non-closed (draft/pending/rejected) version."""
    exists = await bom_svc.check_non_closed_version(db, bom_code, exclude_config_id)
    return success(data={"has_non_closed": exists})


@router.post("/bom-configs/{config_id}/copy", dependencies=[Depends(require_developer)])
async def copy_bom_config(config_id: int, req: BomConfigCopyReq, db: AsyncSession = Depends(get_db)):
    obj = await bom_svc.copy(db, config_id, req.target_bom_code, req.target_bom_name)
    return success(data=BomConfigResp(**obj.to_dict()), message="BOM配置复制成功")


# ── Review / Archive ──

@router.post("/bom-configs/{config_id}/submit-review", dependencies=[Depends(require_developer)])
async def submit_review(
    config_id: int, req: ReviewReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await bom_svc.submit_review(db, config_id, operator=_get_operator(user))
    return success(data=ReviewActionResp(**obj.to_dict()), message="已提交评审")


@router.post("/bom-configs/{config_id}/approve-review", dependencies=[Depends(require_developer)])
async def approve_review(
    config_id: int, req: ReviewReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await bom_svc.approve_review(db, config_id, comment=req.comment, operator=_get_operator(user))
    return success(data=ReviewActionResp(**obj.to_dict()), message="评审已通过")


@router.post("/bom-configs/{config_id}/reject-review", dependencies=[Depends(require_developer)])
async def reject_review(
    config_id: int, req: ReviewReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await bom_svc.reject_review(db, config_id, comment=req.comment, operator=_get_operator(user))
    return success(data=ReviewActionResp(**obj.to_dict()), message="评审已驳回")


@router.post("/bom-configs/{config_id}/withdraw-review", dependencies=[Depends(require_developer)])
async def withdraw_review(
    config_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await bom_svc.withdraw_review(db, config_id, operator=_get_operator(user))
    return success(data=ReviewActionResp(**obj.to_dict()), message="评审已撤回")


@router.post("/bom-configs/{config_id}/archive", dependencies=[Depends(require_developer)])
async def archive_bom(
    config_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await bom_svc.archive_bom(db, config_id, operator=_get_operator(user))
    return success(data=ReviewActionResp(**obj.to_dict()), message="BOM 已归档")


@router.post("/bom-configs/{config_id}/new-iteration", dependencies=[Depends(require_developer)])
async def create_new_iteration(
    config_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await bom_svc.create_new_iteration(db, config_id, operator=_get_operator(user))
    return success(data=BomConfigResp(**obj.to_dict()), message="新迭代版本已创建")


# ── Export Excel ──

@router.post("/bom-configs/{config_id}/export-excel", dependencies=[Depends(require_developer)])
async def export_bom_excel(
    config_id: int,
    db: AsyncSession = Depends(get_db),
):
    from app.services.export_service import export_bom_excel as do_export
    result = await do_export(db, config_id)
    return success(data=result)


# ── Export Diff Report ──

@router.post("/bom-configs/{config_id}/export-diff-report", dependencies=[Depends(require_developer)])
async def export_diff_report(
    config_id: int,
    db: AsyncSession = Depends(get_db),
):
    from app.services.export_service import export_diff_report as do_export
    result = await do_export(db, config_id)
    return success(data=result)


# ── Excel Template / Import / Export ──

@router.post("/bom-configs/{config_id}/export-template", dependencies=[Depends(require_developer)])
async def export_bom_template(
    config_id: int,
    db: AsyncSession = Depends(get_db),
):
    from app.services.export_service import export_bom_template as do_export
    result = await do_export(db, config_id)
    return success(data=result)


@router.post("/bom-configs/{config_id}/export-current", dependencies=[Depends(require_developer)])
async def export_current_config(
    config_id: int,
    db: AsyncSession = Depends(get_db),
):
    from app.services.export_service import export_current_config as do_export
    result = await do_export(db, config_id)
    return success(data=result)


@router.post("/bom-configs/{config_id}/export-pdf", dependencies=[Depends(require_developer)])
async def export_bom_pdf(
    config_id: int,
    db: AsyncSession = Depends(get_db),
):
    from app.services.export_service import export_pdf_report as do_export
    result = await do_export(db, config_id)
    return success(data=result)


@router.post("/bom-configs/{config_id}/import", dependencies=[Depends(require_developer)])
async def import_bom_config(
    config_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        return error(code=400, message="仅支持 .xlsx / .xls 文件")
    content = await file.read()
    from app.services.export_service import import_bom_config as do_import
    result = await do_import(db, config_id, content, operator=_get_operator(user))
    return success(data=result, message="Excel 导入完成")


@router.post("/bom-configs/{config_id}/validate", dependencies=[Depends(require_process)])
async def validate_bom_config(config_id: int, db: AsyncSession = Depends(get_db)):
    """BOM 配置校验 — 必填参数校验，返回校验结果"""
    from app.services.export_service import validate_bom_config as do_validate
    result = await do_validate(db, config_id)
    return success(data=result, message="校验完成")


@router.get("/bom-configs/{config_id}/indicators", dependencies=[Depends(require_process)])
async def list_bom_indicators(config_id: int, db: AsyncSession = Depends(get_db)):
    items = await bom_svc.list_indicators(db, config_id)
    return success(data=[BomIndicatorResp(**i) for i in items])


@router.post("/bom-configs/{config_id}/indicators", dependencies=[Depends(require_developer)])
async def add_bom_indicator(
    config_id: int, req: BomIndicatorCreateReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await bom_svc.add_indicator(db, config_id, req.model_dump(), operator=_get_operator(user))
    items = await bom_svc.list_indicators(db, config_id)
    enriched = next((i for i in items if i["id"] == obj.id), None)
    if enriched is None:
        return success(message="指标添加成功")
    return success(data=BomIndicatorResp(**enriched), message="指标添加成功")


@router.post("/bom-configs/{config_id}/indicators/batch", dependencies=[Depends(require_developer)])
async def batch_add_bom_indicators(
    config_id: int, req: BomIndicatorBatchCreateReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    objs = await bom_svc.batch_add_indicators(db, config_id, [r.model_dump() for r in req.indicators], operator=_get_operator(user))
    items = await bom_svc.list_indicators(db, config_id)
    return success(data=[BomIndicatorResp(**i) for i in items], message="批量导入成功")


@router.put("/bom-configs/{config_id}/indicators/batch-update", dependencies=[Depends(require_developer)])
async def batch_update_bom_indicators(
    config_id: int, req: BomIndicatorBatchUpdateReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    await bom_svc.batch_update_indicators(db, req.ids, req.model_dump(exclude={"ids"}, exclude_none=True), operator=_get_operator(user))
    return success(message="批量更新成功")


@router.put("/bom-configs/{config_id}/indicators/batch-status", dependencies=[Depends(require_developer)])
async def batch_update_indicator_status(
    config_id: int, req: BomIndicatorBatchStatusReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    await bom_svc.batch_update_indicator_status(db, req.ids, req.status, operator=_get_operator(user))
    return success(message="状态更新成功")


@router.put("/bom-configs/indicators/{indicator_id}", dependencies=[Depends(require_developer)])
async def update_bom_indicator(
    indicator_id: int, req: BomIndicatorUpdateReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await bom_svc.update_indicator(db, indicator_id, req.model_dump(exclude_unset=True), operator=_get_operator(user))
    items = await bom_svc.list_indicators(db, obj.bom_config_id)
    enriched = next((i for i in items if i["id"] == obj.id), None)
    if enriched is None:
        return success(message="指标更新成功")
    return success(data=BomIndicatorResp(**enriched), message="指标更新成功")


# ── BomIndicator Actions ──


@router.delete("/bom-configs/indicators/{indicator_id}", dependencies=[Depends(require_developer)])
async def delete_bom_indicator(
    indicator_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    await bom_svc.delete_indicator(db, indicator_id, operator=_get_operator(user))
    return success(message="指标已删除")


# ── Per-param CRUD within a BOM indicator ──

@router.post("/bom-configs/indicators/{bom_indicator_id}/params", dependencies=[Depends(require_developer)])
async def add_bom_indicator_param(
    bom_indicator_id: int, req: BomIndicatorParamAddReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    params = await bom_svc.add_param(db, bom_indicator_id, req.model_dump(), operator=_get_operator(user))
    return success(data=params, message="参数新增成功")


@router.put("/bom-configs/indicators/{bom_indicator_id}/params/{param_key}", dependencies=[Depends(require_developer)])
async def update_bom_indicator_param(
    bom_indicator_id: int, param_key: str, req: BomIndicatorParamUpdateReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    params = await bom_svc.update_param(db, bom_indicator_id, param_key, req.model_dump(exclude_unset=True), operator=_get_operator(user))
    return success(data=params, message="参数更新成功")


@router.delete("/bom-configs/indicators/{bom_indicator_id}/params/{param_key}", dependencies=[Depends(require_developer)])
async def delete_bom_indicator_param(
    bom_indicator_id: int, param_key: str,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    params = await bom_svc.delete_param(db, bom_indicator_id, param_key, operator=_get_operator(user))
    return success(data=params, message="参数已删除")


# ── Collaborative Editing: Batch Save with Optimistic Locking ──
@router.put("/bom-configs/{config_id}/indicators/batch-save", dependencies=[Depends(require_developer)])
async def batch_save_bom_indicator_params(
    config_id: int,
    req: BomIndicatorBatchSaveReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    """协同编辑批量保存：乐观锁并发控制，仅允许编辑自己有权限的项"""
    result = await bom_svc.batch_save_indicator_params(
        db, config_id, [i.model_dump() for i in req.indicators],
        operator_id=user.get("id", 0),
        operator_name=_get_operator(user),
        is_super_admin=user.get("role") == "super_admin",
    )
    return success(data=result, message="保存完成")


# ── Change Logs ──
@router.get("/bom-configs/{config_id}/change-logs", dependencies=[Depends(require_process)])
async def get_bom_change_logs(
    config_id: int,
    test_item_id: int = Query(None),
    indicator_id: int = Query(None),
    db: AsyncSession = Depends(get_db),
):
    logs = await bom_svc.get_change_logs(db, config_id, test_item_id, indicator_id)
    return success(data=[ParamChangeLogResp(**log) for log in logs])


# ── All Indicators Export (per-indicator script) ──


@router.post("/indicators/export", dependencies=[Depends(require_developer)])
async def export_all_indicators(
    req: BomExportReq,
    db: AsyncSession = Depends(get_db),
):
    try:
        result = await ScriptTemplateService.export_all_indicators(db, output_format=req.output_format)
        return success(data=result)
    except Exception as e:
        return error(code=500, message=str(e))


# ── BOM Export (per-indicator script) ──


@router.post("/bom-configs/{config_id}/export", dependencies=[Depends(require_developer)])
async def export_bom_config(
    config_id: int,
    req: BomExportReq,
    db: AsyncSession = Depends(get_db),
):
    try:
        result = await ScriptTemplateService.export_bom_config(
            db, config_id, output_format=req.output_format,
        )
        return success(data=result)
    except Exception as e:
        return error(code=500, message=str(e))


# ════════════════════════════════════════════
# 版本记录
# ════════════════════════════════════════════
@router.get("/versions", dependencies=[Depends(require_process)])
async def list_versions(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    entity_type: str = Query(""),
    entity_id: int = Query(0),
    keyword: str = Query(""),
    operator: str = Query(""),
    date_from: str = Query(""),
    date_to: str = Query(""),
    db: AsyncSession = Depends(get_db),
):
    items, total, p, ps = await ver_svc.list(
        db, page=page, page_size=page_size,
        entity_type=entity_type, entity_id=entity_id,
        keyword=keyword, operator=operator,
        date_from=date_from, date_to=date_to,
    )
    return paginated([VersionSnapshotResp(**i.to_dict()) for i in items], total, p, ps)


@router.get("/versions/{snapshot_id}", dependencies=[Depends(require_process)])
async def get_version_detail(snapshot_id: int, db: AsyncSession = Depends(get_db)):
    obj = await ver_svc.get(db, snapshot_id)
    if not obj:
        from app.core.exceptions import NotFoundError
        raise NotFoundError("版本记录不存在")
    return success(data=VersionSnapshotResp(**obj.to_dict()))


def _diff_params(prev_params: list, cur_params: list) -> dict:
    """Compare two lists of param dicts by 'key' and return structured diff."""
    prev_map = {p["key"]: p for p in prev_params if p.get("key")}
    cur_map = {p["key"]: p for p in cur_params if p.get("key")}
    added = []
    removed = []
    modified = []
    all_keys = set(list(cur_map.keys()) + list(prev_map.keys()))
    for k in sorted(all_keys):
        cur = cur_map.get(k)
        pre = prev_map.get(k)
        if cur and not pre:
            added.append(cur)
        elif pre and not cur:
            removed.append(pre)
        elif cur and pre:
            sub_diff = {}
            for fk in cur:
                if fk == "key":
                    continue
                if str(cur[fk]) != str(pre.get(fk)):
                    sub_diff[fk] = {"before": pre.get(fk), "after": cur[fk]}
            if sub_diff:
                modified.append({"key": k, "name": cur.get("name", k), "diff_fields": sub_diff})
    return {"added": added, "removed": removed, "modified": modified}


@router.get("/versions/{snapshot_id}/diff", dependencies=[Depends(require_process)])
async def diff_version(snapshot_id: int, db: AsyncSession = Depends(get_db)):
    """Compare a snapshot with its immediate predecessor and return diff data."""
    from app.core.exceptions import NotFoundError
    r = await db.execute(select(IndicatorVersionSnapshot).where(IndicatorVersionSnapshot.id == snapshot_id))
    snapshot = r.scalar_one_or_none()
    if not snapshot:
        raise NotFoundError("版本记录不存在")

    prev_r = await db.execute(
        select(IndicatorVersionSnapshot)
        .where(
            IndicatorVersionSnapshot.entity_type == snapshot.entity_type,
            IndicatorVersionSnapshot.entity_id == snapshot.entity_id,
            IndicatorVersionSnapshot.id < snapshot_id,
        )
        .order_by(IndicatorVersionSnapshot.id.desc())
        .limit(1)
    )
    prev = prev_r.scalar_one_or_none()

    current = {}
    previous = {}

    if snapshot.entity_type == "bom":
        for ind in (snapshot.snapshot_data.get("indicators") or []):
            current[ind["indicator_id"]] = ind
        if prev:
            for ind in (prev.snapshot_data.get("indicators") or []):
                previous[ind["indicator_id"]] = ind
    else:
        for item in (snapshot.snapshot_data.get("items") or []):
            current[item["id"]] = item
        if prev:
            for item in (prev.snapshot_data.get("items") or []):
                previous[item["id"]] = item

    added = []
    removed = []
    modified = []
    unchanged = []

    all_keys = set(list(current.keys()) + list(previous.keys()))
    for k in sorted(all_keys):
        cur = current.get(k)
        pre = previous.get(k)
        if cur and not pre:
            added.append(cur)
        elif pre and not cur:
            removed.append(pre)
        elif cur and pre:
            diff_fields = {}
            for field_key in cur:
                if field_key == "id" or field_key.endswith("_id"):
                    continue
                if str(cur[field_key]) != str(pre.get(field_key)):
                    if field_key == "params":
                        diff_fields[field_key] = _diff_params(pre.get(field_key, []), cur[field_key])
                    else:
                        diff_fields[field_key] = {"before": pre.get(field_key), "after": cur[field_key]}
            if diff_fields:
                modified.append({"item": cur, "diff_fields": diff_fields})
            else:
                unchanged.append(cur)

    return success(data={
        "snapshot_id": snapshot_id,
        "entity_type": snapshot.entity_type,
        "entity_id": snapshot.entity_id,
        "version": snapshot.version,
        "prev_version": prev.version if prev else None,
        "change_summary": snapshot.change_summary,
        "operator": snapshot.operator,
        "created_at": snapshot.created_at.isoformat() if snapshot.created_at else None,
        "diff": {
            "added": added,
            "removed": removed,
            "modified": modified,
            "total_added": len(added),
            "total_removed": len(removed),
            "total_modified": len(modified),
            "total_unchanged": len(unchanged),
        },
    })


@router.get("/versions/diff", dependencies=[Depends(require_process)])
async def diff_two_versions(
    entity_type: str = Query(..., description="实体类型: bom 或 collection"),
    entity_id: int = Query(..., description="实体ID"),
    v1_id: int = Query(..., description="版本1快照ID"),
    v2_id: int = Query(..., description="版本2快照ID"),
    db: AsyncSession = Depends(get_db),
):
    """Compare any two version snapshots and return detailed diff including parameter differences."""
    from app.core.exceptions import NotFoundError
    
    r1 = await db.execute(select(IndicatorVersionSnapshot).where(IndicatorVersionSnapshot.id == v1_id))
    snapshot1 = r1.scalar_one_or_none()
    r2 = await db.execute(select(IndicatorVersionSnapshot).where(IndicatorVersionSnapshot.id == v2_id))
    snapshot2 = r2.scalar_one_or_none()
    
    if not snapshot1 or not snapshot2:
        raise NotFoundError("版本快照不存在")
    if snapshot1.entity_type != entity_type or snapshot2.entity_type != entity_type:
        raise ValueError("快照实体类型不匹配")
    if snapshot1.entity_id != entity_id or snapshot2.entity_id != entity_id:
        raise ValueError("快照实体ID不匹配")
    
    current = {}
    previous = {}
    
    if entity_type == "bom":
        for ind in (snapshot1.snapshot_data.get("indicators") or []):
            current[ind["indicator_id"]] = ind
        for ind in (snapshot2.snapshot_data.get("indicators") or []):
            previous[ind["indicator_id"]] = ind
    else:
        for item in (snapshot1.snapshot_data.get("items") or []):
            current[item["id"]] = item
        for item in (snapshot2.snapshot_data.get("items") or []):
            previous[item["id"]] = item
    
    added = []
    removed = []
    modified = []
    unchanged = []
    
    all_keys = set(list(current.keys()) + list(previous.keys()))
    for k in sorted(all_keys):
        cur = current.get(k)
        pre = previous.get(k)
        if cur and not pre:
            added.append(cur)
        elif pre and not cur:
            removed.append(pre)
        elif cur and pre:
            diff_fields = {}
            for field_key in cur:
                if field_key == "id" or field_key.endswith("_id"):
                    continue
                if str(cur[field_key]) != str(pre.get(field_key)):
                    if field_key == "params":
                        diff_fields[field_key] = _diff_params(pre.get(field_key, []), cur[field_key])
                    else:
                        diff_fields[field_key] = {"before": pre.get(field_key), "after": cur[field_key]}
            if diff_fields:
                modified.append({"item": cur, "diff_fields": diff_fields})
            else:
                unchanged.append(cur)
    
    return success(data={
        "v1_snapshot_id": v1_id,
        "v2_snapshot_id": v2_id,
        "entity_type": entity_type,
        "entity_id": entity_id,
        "v1_version": snapshot1.version,
        "v2_version": snapshot2.version,
        "diff": {
            "added": added,
            "removed": removed,
            "modified": modified,
            "total_added": len(added),
            "total_removed": len(removed),
            "total_modified": len(modified),
            "total_unchanged": len(unchanged),
        },
    })


@router.get("/bom-configs/{config_id}/diff-baseline", dependencies=[Depends(require_process)])
async def diff_baseline(
    config_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Compare current unpublished BOM with latest published/archived version (baseline)."""
    from app.core.exceptions import NotFoundError
    
    config = await bom_svc.get(db, config_id)
    if not config:
        raise NotFoundError("BOM配置不存在")
    
    # Get latest approved/archived snapshot for this BOM
    r = await db.execute(
        select(IndicatorVersionSnapshot)
        .where(
            IndicatorVersionSnapshot.entity_type == "bom",
            IndicatorVersionSnapshot.entity_id == config_id,
            IndicatorVersionSnapshot.change_summary.ilike("%评审通过%") | 
            IndicatorVersionSnapshot.change_summary.ilike("%归档%")
        )
        .order_by(IndicatorVersionSnapshot.id.desc())
        .limit(1)
    )
    baseline = r.scalar_one_or_none()
    
    if not baseline:
        return success(data={
            "has_baseline": False,
            "message": "无已发布或归档版本作为基准",
            "diff": {"added": [], "removed": [], "modified": [], "total_added": 0, "total_removed": 0, "total_modified": 0},
        })
    
    # Get current BOM indicators (unpublished changes)
    indicators = await bom_svc.list_indicators(db, config_id)
    current = {}
    for ind in indicators:
        current[ind["indicator_id"]] = ind
    
    # Get baseline indicators
    previous = {}
    for ind in (baseline.snapshot_data.get("indicators") or []):
        previous[ind["indicator_id"]] = ind
    
    added = []
    removed = []
    modified = []
    unchanged = []
    
    all_keys = set(list(current.keys()) + list(previous.keys()))
    for k in sorted(all_keys):
        cur = current.get(k)
        pre = previous.get(k)
        if cur and not pre:
            added.append(cur)
        elif pre and not cur:
            removed.append(pre)
        elif cur and pre:
            diff_fields = {}
            for field_key in cur:
                if field_key == "id" or field_key.endswith("_id"):
                    continue
                if str(cur[field_key]) != str(pre.get(field_key)):
                    if field_key == "params":
                        diff_fields[field_key] = _diff_params(pre.get(field_key, []), cur[field_key])
                    else:
                        diff_fields[field_key] = {"before": pre.get(field_key), "after": cur[field_key]}
            if diff_fields:
                modified.append({"item": cur, "diff_fields": diff_fields})
            else:
                unchanged.append(cur)
    
    return success(data={
        "has_baseline": True,
        "baseline_snapshot_id": baseline.id,
        "baseline_version": baseline.version,
        "diff": {
            "added": added,
            "removed": removed,
            "modified": modified,
            "total_added": len(added),
            "total_removed": len(removed),
            "total_modified": len(modified),
            "total_unchanged": len(unchanged),
        },
    })


@router.post("/versions/{snapshot_id}/rollback", dependencies=[Depends(require_developer)])
async def rollback_version(
    snapshot_id: int, req: RollbackReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    operator = req.operator or _get_operator(user)
    snapshot = await ver_svc.rollback(db, snapshot_id, operator)
    return success(data=VersionSnapshotResp(**snapshot.to_dict()), message="回滚成功")


# ════════════════════════════════════════════
# 联合查询 & 导出
# ════════════════════════════════════════════
@router.get("/query", dependencies=[Depends(require_process)])
async def query_indicators(
    bom_code: str = Query(""),
    collection_id: int = Query(None),
    indicator_name: str = Query(""),
    product_type: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(
        BomConfig.id.label("bom_config_id"),
        BomConfig.bom_code,
        BomConfig.bom_name,
        TestItemCollection.id.label("collection_id"),
        TestItemCollection.name.label("collection_name"),
        IndicatorDict.id.label("indicator_id"),
        IndicatorDict.code.label("indicator_code"),
        IndicatorDict.name.label("indicator_name"),
        IndicatorDict.category,
        IndicatorDict.params,
        BomIndicator.unit,
        BomIndicator.judgment_rule,
        BomIndicator.test_stage,
    ).select_from(BomConfig) \
        .join(TestItemCollection, BomConfig.collection_id == TestItemCollection.id) \
        .join(BomIndicator, BomConfig.id == BomIndicator.bom_config_id) \
        .join(IndicatorDict, BomIndicator.indicator_id == IndicatorDict.id)
    if bom_code:
        stmt = stmt.where(BomConfig.bom_code.ilike(f"%{bom_code}%"))
    if collection_id:
        stmt = stmt.where(TestItemCollection.id == collection_id)
    if indicator_name:
        stmt = stmt.where(IndicatorDict.name.ilike(f"%{indicator_name}%"))
    if product_type:
        stmt = stmt.where(TestItemCollection.product_type.ilike(f"%{product_type}%"))

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total_result = await db.execute(count_stmt)
    total = total_result.scalar() or 0

    import math
    total_pages = max(1, math.ceil(total / page_size))
    offset = (page - 1) * page_size
    r = await db.execute(stmt.offset(offset).limit(page_size))
    rows = []
    for row in r.all():
        d = row._asdict()
        rows.append(IndicatorQueryResp(**d))
    return paginated(
        [r.model_dump() for r in rows],
        total, page, page_size,
    )


@router.post("/export", dependencies=[Depends(require_process)])
async def export_indicators(
    bom_code: str = Query(""),
    collection_id: int = Query(None),
    indicator_name: str = Query(""),
    product_type: str = Query(""),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(
        BomConfig.bom_code,
        BomConfig.bom_name,
        BomConfig.collection_version,
        TestItemCollection.name.label("collection_name"),
        TestItemCollection.version.label("collection_current_version"),
        CollectionTestItem.name.label("test_item_name"),
        CollectionTestItem.service_address,
        CollectionTestItem.timeout_seconds,
        CollectionTestItem.block_type,
        CollectionTestItem.parallel_enabled,
        IndicatorDict.code.label("indicator_code"),
        IndicatorDict.name.label("indicator_name"),
        IndicatorDict.category,
        IndicatorDict.params,
        BomIndicator.unit,
        BomIndicator.judgment_rule,
        BomIndicator.test_stage,
    ).select_from(BomConfig) \
        .join(TestItemCollection, BomConfig.collection_id == TestItemCollection.id) \
        .outerjoin(CollectionTestItem, TestItemCollection.id == CollectionTestItem.collection_id) \
        .join(BomIndicator, BomConfig.id == BomIndicator.bom_config_id) \
        .join(IndicatorDict, BomIndicator.indicator_id == IndicatorDict.id) \
        .where(BomIndicator.status == 1)
    if bom_code:
        stmt = stmt.where(BomConfig.bom_code.ilike(f"%{bom_code}%"))
    if collection_id:
        stmt = stmt.where(TestItemCollection.id == collection_id)
    if indicator_name:
        stmt = stmt.where(IndicatorDict.name.ilike(f"%{indicator_name}%"))
    if product_type:
        stmt = stmt.where(TestItemCollection.product_type.ilike(f"%{product_type}%"))
    stmt = stmt.order_by(BomConfig.bom_code, TestItemCollection.name, CollectionTestItem.sort_order)

    r = await db.execute(stmt)
    headers = ["BOM编码", "BOM名称", "绑定集合", "集合版本", "是否最新", "测试项名称", "微服务地址", "超时时间(秒)", "阻断类型", "支持并行", "指标编码", "指标名称", "分类", "硬件参数", "单位", "判定规则", "测试阶段"]
    block_map = {"must_test": "必测不可屏蔽", "critical": "关键阻断项", "normal": "普通项"}
    rows = []
    for row in r.all():
        is_latest = row.collection_version == row.collection_current_version
        params_str = "; ".join(f"{k}={v}" for k, v in (row.params or {}).items()) if isinstance(row.params, dict) else str(row.params or "")
        rows.append([
            row.bom_code, row.bom_name, row.collection_name,
            row.collection_version,
            "是" if is_latest else "否",
            row.test_item_name, row.service_address or "",
            row.timeout_seconds or "",
            block_map.get(row.block_type, row.block_type or ""),
            "是" if row.parallel_enabled else "否",
            row.indicator_code, row.indicator_name, row.category,
            params_str, row.unit,
            row.judgment_rule, row.test_stage,
        ])
    xlsx_bytes = export_xlsx(headers, rows)
    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=indicators.xlsx"},
    )


# ── Script Template ──


@router.get("/script-templates", dependencies=[Depends(require_process)])
async def list_script_templates(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: str = Query(""),
    status: int = Query(None),
    db: AsyncSession = Depends(get_db),
):
    data = await ScriptTemplateService.list(db, page=page, page_size=page_size, keyword=keyword, status=status)
    items, total, p, ps = data
    return paginated([ScriptTemplateResp(**i.to_dict()) for i in items], total, p, ps)


@router.get("/script-templates/active", dependencies=[Depends(require_process)])
async def list_active_scripts(db: AsyncSession = Depends(get_db)):
    items = await ScriptTemplateService.list_active(db)
    return success(data=[ScriptTemplateResp(**i.to_dict()) for i in items])


@router.get("/script-templates/{script_id}", dependencies=[Depends(require_process)])
async def get_script_template(script_id: int, db: AsyncSession = Depends(get_db)):
    obj = await ScriptTemplateService.get(db, script_id)
    return success(data=ScriptTemplateResp(**obj.to_dict()))


@router.post("/script-templates", dependencies=[Depends(require_developer)])
async def create_script_template(
    req: ScriptTemplateCreateReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await ScriptTemplateService.create(db, req.model_dump(), operator=user.get("username", ""))
    return success(data=ScriptTemplateResp(**obj.to_dict()), message="脚本模板创建成功")


@router.put("/script-templates/{script_id}", dependencies=[Depends(require_developer)])
async def update_script_template(
    script_id: int,
    req: ScriptTemplateUpdateReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await ScriptTemplateService.update(db, script_id, req.model_dump(exclude_none=True), operator=user.get("username", ""))
    return success(data=ScriptTemplateResp(**obj.to_dict()), message="脚本模板更新成功")


@router.delete("/script-templates/{script_id}", dependencies=[Depends(require_developer)])
async def delete_script_template(script_id: int, db: AsyncSession = Depends(get_db)):
    await ScriptTemplateService.delete(db, script_id)
    return success(message="脚本模板已删除")


@router.put("/script-templates/{script_id}/status", dependencies=[Depends(require_developer)])
async def toggle_script_status(
    script_id: int,
    status: int = Query(..., description="1启用 0禁用"),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    obj = await ScriptTemplateService.toggle_status(db, script_id, status, operator=user.get("username", ""))
    return success(data=ScriptTemplateResp(**obj.to_dict()), message="状态已更新")


@router.get("/script-exports/{file_name}")
async def download_script_export(file_name: str):
    from app.config import get_settings
    settings = get_settings()
    file_path = os.path.join(settings.UPLOAD_FOLDER, "script_exports", file_name)
    if not os.path.isfile(file_path):
        from fastapi.responses import JSONResponse
        return JSONResponse({"code": 404, "message": "文件不存在或已过期"}, status_code=404)
    return StreamingResponse(
        open(file_path, "rb"),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(file_name)}"},
    )


@router.get("/excel-exports/{file_name}")
async def download_excel_export(file_name: str):
    from app.config import get_settings
    settings = get_settings()
    file_path = os.path.join(settings.UPLOAD_FOLDER, "excel_exports", file_name)
    if not os.path.isfile(file_path):
        from fastapi.responses import JSONResponse
        return JSONResponse({"code": 404, "message": "文件不存在或已过期"}, status_code=404)
    return StreamingResponse(
        open(file_path, "rb"),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(file_name, safe='')}"},
    )


@router.post("/script-templates/validate", dependencies=[Depends(require_developer)])
async def validate_script_source(data: dict = Body(...), db: AsyncSession = Depends(get_db)):
    """Validate Python syntax of script source code."""
    source = data.get("source_code", "")
    if not source:
        return error(code=400, message="代码为空")
    import ast
    try:
        ast.parse(source)
        return success(data={"valid": True, "message": "语法正确"})
    except SyntaxError as e:
        return success(data={
            "valid": False,
            "message": f"第 {e.lineno} 行: {e.msg}",
            "lineno": e.lineno,
            "offset": e.offset,
            "text": e.text,
        })


@router.post("/script-templates/execute", dependencies=[Depends(require_developer)])
async def execute_script(
    req: ScriptExecuteReq,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    try:
        result = await ScriptTemplateService.execute_script(
            db, req.script_id,
            indicator_ids=req.indicator_ids or None,
            collection_ids=req.collection_ids or None,
            bom_config_ids=req.bom_config_ids or None,
            export_all=req.export_all,
            operator=user.get("username", ""),
        )
        return success(data=result)
    except ValueError as e:
        return error(code=400, message=str(e))
    except RuntimeError as e:
        return error(code=500, message=str(e))
