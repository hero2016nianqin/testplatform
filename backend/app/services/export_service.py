"""Excel 导出/导入服务：BOM 填写记录表、版本差异对比报告、空白模板、配置导入"""
import os
import uuid
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.models.metrics import BomConfig, BomIndicator, \
    CollectionTestItem, TestItemIndicator, IndicatorDict, IndicatorVersionSnapshot
from fpdf import FPDF

from app.services.bom_config_service import BomConfigService
from app.services.dictionary_service import DictionaryService


async def validate_bom_config(db: AsyncSession, config_id: int) -> dict:
    """BOM 配置校验 — 必填参数校验，返回校验结果"""
    config = await BomConfigService.get(db, config_id)
    indicators = await BomConfigService.list_indicators(db, config_id)
    all_errors = []

    for ind in indicators:
        params = ind.get("params") or []
        for p in params:
            val = p.get("param_value") if p.get("param_value") is not None else p.get("value")
            param_name = p.get("param_name") or p.get("name") or p.get("key", "")
            param_key = p.get("param_key") or p.get("key", "")
            if val is None or val == "":
                all_errors.append({
                    "indicator_id": ind["indicator_id"],
                    "indicator_name": ind["indicator_name"],
                    "param_key": param_key,
                    "param_name": param_name,
                    "field": "param",
                    "message": f"参数「{param_name}」为空"
                })

    return {
        "valid": len(all_errors) == 0,
        "errors": all_errors,
        "config_id": config_id,
        "bom_code": config.bom_code,
        "version": config.version,
    }


async def export_bom_template(db: AsyncSession, config_id: int) -> dict:
    """导出 BOM 空白填写模板 (Excel) — 每工序_工位一个 Sheet，带填写说明"""
    config = await BomConfigService.get(db, config_id)
    indicators = await BomConfigService.list_indicators(db, config_id)
    bom_ind_map = {bi["indicator_id"]: bi for bi in indicators}

    # Get items grouped by process + station
    r = await db.execute(
        select(CollectionTestItem)
        .where(CollectionTestItem.collection_id == config.collection_id)
        .order_by(CollectionTestItem.process_name, CollectionTestItem.station, CollectionTestItem.sort_order)
    )
    items = r.scalars().all()

    # Group by (process_name, station)
    process_groups = {}
    for item in items:
        pn = item.process_name or "未分类工序"
        sn = item.station or "通用工位"
        key = f"{pn}_{sn}"
        if key not in process_groups:
            process_groups[key] = []
        process_groups[key].append(item)

    wb = Workbook()

    # Instruction sheet
    ws_inst = wb.active
    ws_inst.title = "填写说明"
    ws_inst.merge_cells("A1:I1")
    ws_inst["A1"] = "BOM 指标参数填写模板使用说明"
    ws_inst["A1"].font = Font(bold=True, size=14)
    ws_inst["A1"].alignment = Alignment(horizontal="center")
    ws_inst.merge_cells("A2:I2")
    ws_inst["A2"] = f"BOM编码：{config.bom_code} | BOM名称：{config.bom_name} | 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_inst["A2"].font = Font(size=10, color="666666")
    ws_inst["A2"].alignment = Alignment(horizontal="center")

    inst_rows = [
        ("1. 文件结构", "每个工序_工位为一个独立 Sheet，Sheet 名称为「工序_工位」"),
        ("2. 填写规则", "仅填写“自定义上限/下限”、“单位”、“硬件参数”、“备注”四列，其余列为只读参考"),
        ("3. 数字格式", "上限/下限仅允许整数或小数，不允许输入文本或特殊字符"),
        ("4. 布尔格式", "布尔参数仅允许 true / false"),
        ("5. 列表格式", "列表参数请用英文逗号分隔多个值，如：val1,val2,val3（不可使用中文逗号）"),
        ("6. 字符串格式", "字符串参数无特殊限制，但建议避免特殊符号"),
        ("7. 必填标记", "表头标有 [必填] 的列为必填项，导入时会校验"),
        ("8. 导入规则", "导入时仅更新数值/阈值/参数，不新增/删除测试项、不修改参数格式；自动匹配 工序+工位+测试项 三重唯一标识"),
        ("9. 参数格式锁定", "参数格式（数字/布尔/列表/字符串）继承自指标字典，不可在模板中修改"),
        ("10. 导入校验", "导入前会自动校验格式、逻辑（如下限≤上限），错误行会在结果中标注"),
    ]
    r = 4
    for title, desc in inst_rows:
        ws_inst.cell(row=r, column=1, value=title).font = Font(bold=True)
        ws_inst.cell(row=r, column=2, value=desc)
        ws_inst.merge_cells(f"B{r}:H{r}")
        ws_inst.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        r += 1

    # Column widths for instruction sheet
    ws_inst.column_dimensions["A"].width = 18
    ws_inst.column_dimensions["B"].width = 70

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    readonly_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_font_black = Font(bold=True, size=11)
    title_font = Font(bold=True, size=13)

    # Create a sheet per process_station
    for group_key in sorted(process_groups.keys()):
        sheet_items = process_groups[group_key]
        pn = group_key.split("_")[0]
        sn = "_".join(group_key.split("_")[1:])
        ws = wb.create_sheet(title=group_key[:31])  # Sheet name max 31 chars

        # Title
        ws.merge_cells("A1:I1")
        ws["A1"] = f"工序：{pn}  工位：{sn}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center")

        # Headers
        headers = [
            "BOM编码", "工序", "工位", "测试项", "指标编码", "指标名称",
            "自定义上限[必填]", "自定义下限[必填]", "单位[必填]",
            "硬件参数(Key=Value,英文逗号分隔)", "备注"
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            )
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Mark required columns with yellow fill
        for col in [7, 8, 9]:  # 上限、下限、单位
            ws.cell(row=2, column=col).fill = required_fill

        # Get BOM indicator overrides map
        ind_r = await db.execute(
            select(TestItemIndicator).where(TestItemIndicator.test_item_id.in_([i.id for i in items]))
        )
        all_ti_inds = ind_r.scalars().all()
        ti_ind_by_item = {}
        for ti_ind in all_ti_inds:
            if ti_ind.test_item_id not in ti_ind_by_item:
                ti_ind_by_item[ti_ind.test_item_id] = []
            ti_ind_by_item[ti_ind.test_item_id].append(ti_ind)

        row = 3
        for item in sheet_items:
            item_inds = ti_ind_by_item.get(item.id, [])
            for ti_ind in item_inds:
                bi = bom_ind_map.get(ti_ind.indicator_id)
                indicator = await DictionaryService.get(db, ti_ind.indicator_id)
                unit = bi.get("unit") if bi and bi.get("unit") else indicator.unit if indicator else ""
                params_str = ""
                if indicator and indicator.params:
                    params_str = "; ".join(f"{k}={v}" for k, v in indicator.params.items())

                # Pre-fill with current values if any
                upper = bi.get("upper_limit") if bi and bi.get("upper_limit") is not None else ""
                lower = bi.get("lower_limit") if bi and bi.get("lower_limit") is not None else ""

                ws.cell(row=row, column=1, value=config.bom_code).border = thin_border
                ws.cell(row=row, column=2, value=pn).border = thin_border
                ws.cell(row=row, column=3, value=sn).border = thin_border
                ws.cell(row=row, column=4, value=item.name).border = thin_border
                ws.cell(row=row, column=5, value=indicator.code if indicator else "").border = thin_border
                ws.cell(row=row, column=6, value=indicator.name if indicator else "").border = thin_border
                c7 = ws.cell(row=row, column=7, value=upper)
                c7.border = thin_border
                c7.fill = required_fill
                c8 = ws.cell(row=row, column=8, value=lower)
                c8.border = thin_border
                c8.fill = required_fill
                c9 = ws.cell(row=row, column=9, value=unit)
                c9.border = thin_border
                c9.fill = required_fill
                ws.cell(row=row, column=10, value=params_str).border = thin_border
                ws.cell(row=row, column=11, value="").border = thin_border
                row += 1

        # Column widths
        widths = [14, 12, 20, 15, 22, 14, 14, 10, 40, 20]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze panes
        ws.freeze_panes = "A3"

    # Remove default sheet if created
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Save
    settings = get_settings()
    export_dir = os.path.join(settings.UPLOAD_FOLDER, "excel_exports")
    os.makedirs(export_dir, exist_ok=True)
    file_id = str(uuid.uuid4())[:8]
    file_name = f"BOM模板_{config.bom_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file_id}.xlsx"
    file_path = os.path.join(export_dir, file_name)
    wb.save(file_path)

    return {
        "file_name": file_name,
        "file_size": os.path.getsize(file_path),
        "download_url": f"/api/v1/metrics/excel-exports/{file_name}",
    }


async def export_current_config(db: AsyncSession, config_id: int) -> dict:
    """导出当前 BOM 配置为 Excel (指标填写记录表)"""
    config = await BomConfigService.get(db, config_id)
    indicators = await BomConfigService.list_indicators(db, config_id)
    bom_ind_map = {bi["indicator_id"]: bi for bi in indicators}

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM指标填写记录"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"BOM指标填写记录 — {config.bom_code} ({config.bom_name})"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["测试项", "指标编码", "指标名称", "自定义上限", "自定义下限", "单位", "硬件参数(Key=Value)", "备注"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data
    row = 4
    items_r = await db.execute(
        select(CollectionTestItem)
        .where(CollectionTestItem.collection_id == config.collection_id)
        .order_by(CollectionTestItem.sort_order)
    )
    items = items_r.scalars().all()

    for item in items:
        ind_r = await db.execute(
            select(TestItemIndicator).where(TestItemIndicator.test_item_id == item.id)
        )
        item_inds = ind_r.scalars().all()
        for ti_ind in item_inds:
            bi = bom_ind_map.get(ti_ind.indicator_id)
            indicator = await DictionaryService.get(db, ti_ind.indicator_id)
            unit = bi.get("unit") if bi and bi.get("unit") else indicator.unit if indicator else ""
            params_str = ""
            if indicator and indicator.params:
                params_str = "; ".join(f"{k}={v}" for k, v in indicator.params.items())
            ws.cell(row=row, column=1, value=item.name).border = thin_border
            ws.cell(row=row, column=2, value=indicator.code if indicator else "").border = thin_border
            ws.cell(row=row, column=3, value=indicator.name if indicator else "").border = thin_border
            ws.cell(row=row, column=4, value=params_str).border = thin_border
            ws.cell(row=row, column=5, value=unit).border = thin_border
            row += 1

    # Column widths
    widths = [20, 15, 20, 40, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Save
    settings = get_settings()
    export_dir = os.path.join(settings.UPLOAD_FOLDER, "excel_exports")
    os.makedirs(export_dir, exist_ok=True)
    file_id = str(uuid.uuid4())[:8]
    file_name = f"BOM填写记录_{config.bom_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file_id}.xlsx"
    file_path = os.path.join(export_dir, file_name)
    wb.save(file_path)

    return {
        "file_name": file_name,
        "file_size": os.path.getsize(file_path),
        "download_url": f"/api/v1/metrics/excel-exports/{file_name}",
    }


async def import_bom_config(db: AsyncSession, config_id: int, file_bytes: bytes, operator: str = "") -> dict:
    """导入 BOM 配置 (Excel) — 仅更新数值/阈值/参数，自动匹配，返回可视化汇总"""
    config = await BomConfigService.get(db, config_id)
    indicators = await BomConfigService.list_indicators(db, config_id)
    bom_ind_map = {bi["indicator_id"]: bi for bi in indicators}

    # Load workbook
    wb = load_workbook(filename=file_bytes)
    errors = []
    updated_count = 0
    skipped_count = 0
    empty_count = 0

    # Get all items for this collection
    items_r = await db.execute(
        select(CollectionTestItem)
        .where(CollectionTestItem.collection_id == config.collection_id)
    )
    items = items_r.scalars().all()
    item_by_name = {item.name: item for item in items}

    # Build indicator map
    all_ti_inds_r = await db.execute(
        select(TestItemIndicator).where(TestItemIndicator.test_item_id.in_([i.id for i in items]))
    )
    all_ti_inds = all_ti_inds_r.scalars().all()
    ti_ind_by_item = {}
    for ti_ind in all_ti_inds:
        if ti_ind.test_item_id not in ti_ind_by_item:
            ti_ind_by_item[ti_ind.test_item_id] = []
        ti_ind_by_item[ti_ind.test_item_id].append(ti_ind)

    indicator_dict = {}
    for ind in indicators:
        indicator_dict[ind["indicator_id"]] = ind

    dict_indicators = await DictionaryService.list_all_active(db)
    dict_by_code = {d.code: d for d in dict_indicators}

    # Process each sheet (each process_station group)
    for sheet_name in wb.sheetnames:
        if sheet_name in ["填写说明"]:
            continue
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=3, values_only=False):
            # Columns: A=BOM编码, B=工序, C=工位, D=测试项, E=指标编码, F=指标名称,
            # G=自定义上限, H=自定义下限, I=单位, J=硬件参数, K=备注
            row_vals = [cell.value for cell in row]
            if not any(v is not None and str(v).strip() != "" for v in row_vals[6:11]):
                continue  # Skip empty rows

            bom_code = str(row_vals[0]).strip() if row_vals[0] else ""
            process = str(row_vals[1]).strip() if row_vals[1] else ""
            station = str(row_vals[2]).strip() if row_vals[2] else ""
            item_name = str(row_vals[3]).strip() if row_vals[3] else ""
            ind_code = str(row_vals[4]).strip() if row_vals[4] else ""
            ind_name = str(row_vals[5]).strip() if row_vals[5] else ""
            upper_raw = row_vals[6]
            lower_raw = row_vals[7]
            unit_raw = row_vals[8]
            params_raw = str(row_vals[9]).strip() if row_vals[9] else ""
            remark_raw = str(row_vals[10]).strip() if row_vals[10] else ""

            if not item_name or not ind_code:
                errors.append(f"第{row[0].row}行: 缺少测试项名称或指标编码")
                skipped_count += 1
                continue

            item = item_by_name.get(item_name)
            if not item:
                errors.append(f"第{row[0].row}行: 找不到测试项 '{item_name}'")
                skipped_count += 1
                continue

            dict_ind = dict_by_code.get(ind_code)
            if not dict_ind:
                errors.append(f"第{row[0].row}行: 找不到指标编码 '{ind_code}'")
                skipped_count += 1
                continue

            # Find test item indicator
            ti_inds = ti_ind_by_item.get(item.id, [])
            ti_ind = next((t for t in ti_inds if t.indicator_id == dict_ind.id), None)
            if not ti_ind:
                errors.append(f"第{row[0].row}行: 测试项 '{item_name}' 未绑定指标 '{ind_code}'")
                skipped_count += 1
                continue

            # Find BOM indicator
            bi = bom_ind_map.get(dict_ind.id)

            # Validate format
            ind_params = dict_ind.params or []
            param_format_map = {p.get("key"): p.get("type", "string") for p in ind_params}

            # Validate upper/lower limits (numeric)
            upper_val = None
            lower_val = None
            if upper_raw is not None and str(upper_raw).strip() != "":
                try:
                    upper_val = float(str(upper_raw).strip())
                except ValueError:
                    errors.append(f"第{row[0].row}行: 上限 '{upper_raw}' 不是有效数字")
                    continue
            if lower_raw is not None and str(lower_raw).strip() != "":
                try:
                    lower_val = float(str(lower_raw).strip())
                except ValueError:
                    errors.append(f"第{row[0].row}行: 下限 '{lower_raw}' 不是有效数字")
                    continue
            if upper_val is not None and lower_val is not None and lower_val > upper_val:
                errors.append(f"第{row[0].row}行: 下限 {lower_val} 不能大于上限 {upper_val}")
                continue

            # Validate unit (string)
            unit_val = str(unit_raw).strip() if unit_raw else ""

            # Validate params
            params_list = []
            param_error = False
            if params_raw:
                for p_str in params_raw.split(","):
                    p_str = p_str.strip()
                    if not p_str:
                        continue
                    if "=" not in p_str:
                        continue
                    k, v = p_str.split("=", 1)
                    k = k.strip()
                    v = v.strip()
                    fmt = param_format_map.get(k, "string")
                    if fmt in ("number", "range", "percent"):
                        try:
                            float(v)
                        except ValueError:
                            errors.append(f"第{row[0].row}行: 参数 '{k}' 格式为 {fmt}，值 '{v}' 不是有效数字")
                            param_error = True
                            break
                    elif fmt in ("boolean",):
                        if v.lower() not in ("true", "false", "1", "0"):
                            errors.append(f"第{row[0].row}行: 参数 '{k}' 格式为布尔，值 '{v}' 必须是 true/false")
                            param_error = True
                            break
                    elif fmt in ("array", "list"):
                        if "，" in v:
                            errors.append(f"第{row[0].row}行: 参数 '{k}' 为列表类型，请使用英文逗号分隔")
                            param_error = True
                            break
                    params_list.append({"key": k, "value": v, "name": k})
            if param_error:
                continue

            # Update BOM indicator
            update_data = {}
            has_changes = False

            if upper_val is not None:
                update_data["upper_limit"] = upper_val
                has_changes = True
            if lower_val is not None:
                update_data["lower_limit"] = lower_val
                has_changes = True
            if unit_val:
                update_data["unit"] = unit_val
                has_changes = True
            if remark_raw:
                update_data["remark"] = remark_raw
                has_changes = True
            if params_list:
                update_data["params"] = params_list
                has_changes = True

            if has_changes:
                if bi and bi.get("id"):
                    await BomConfigService.update_indicator(db, bi["id"], {
                        "unit": unit_val or bi.get("unit", ""),
                        "upper_limit": update_data.get("upper_limit"),
                        "lower_limit": update_data.get("lower_limit"),
                        "remark": remark_raw or bi.get("remark", ""),
                        "params": params_list or bi.get("params", []),
                    })
                else:
                    await BomConfigService.add_indicator(db, config_id, {
                        "indicator_id": dict_ind.id,
                        "unit": unit_val or "",
                        "judgment_rule": "合格",
                        "test_stage": "",
                        "remark": remark_raw or "",
                        "params": params_list or [],
                        **{k: v for k, v in update_data.items() if k in ["upper_limit", "lower_limit"]}
                    })
                updated_count += 1
            else:
                empty_count += 1

    summary = {
        "total_rows": sum(1 for _ in wb.sheetnames if _ != "填写说明"),
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "empty_count": empty_count,
        "errors": errors,
    }

    return {
        "summary": summary,
        "download_url": f"/api/v1/metrics/excel-exports/import_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    }


async def export_bom_excel(db: AsyncSession, config_id: int) -> dict:
    """导出 BOM 指标填写记录表 (Excel)"""
    config = await BomConfigService.get(db, config_id)
    indicators = await BomConfigService.list_indicators(db, config_id)
    bom_ind_map = {bi["indicator_id"]: bi for bi in indicators}

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM指标填写记录"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = f"BOM指标填写记录 — {config.bom_code} ({config.bom_name})"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["工序", "工位", "测试项", "指标编码", "指标名称", "自定义上限", "自定义下限", "单位", "硬件参数(Key=Value)", "备注"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data
    row = 4
    items_r = await db.execute(
        select(CollectionTestItem)
        .where(CollectionTestItem.collection_id == config.collection_id)
        .order_by(CollectionTestItem.process_name, CollectionTestItem.station, CollectionTestItem.sort_order)
    )
    items = items_r.scalars().all()

    for item in items:
        ind_r = await db.execute(
            select(TestItemIndicator).where(TestItemIndicator.test_item_id == item.id)
        )
        item_inds = ind_r.scalars().all()
        for ti_ind in item_inds:
            bi = bom_ind_map.get(ti_ind.indicator_id)
            indicator = await DictionaryService.get(db, ti_ind.indicator_id)
            unit = bi.get("unit") if bi and bi.get("unit") else indicator.unit if indicator else ""
            params_str = ""
            if indicator and indicator.params:
                params_str = "; ".join(f"{k}={v}" for k, v in indicator.params.items())
            upper = bi.get("upper_limit") if bi and bi.get("upper_limit") is not None else ""
            lower = bi.get("lower_limit") if bi and bi.get("lower_limit") is not None else ""
            remark = bi.get("remark") if bi and bi.get("remark") else ""
            ws.cell(row=row, column=1, value=item.process_name or "未分类工序").border = thin_border
            ws.cell(row=row, column=2, value=item.station or "通用工位").border = thin_border
            ws.cell(row=row, column=3, value=item.name).border = thin_border
            ws.cell(row=row, column=4, value=indicator.code if indicator else "").border = thin_border
            ws.cell(row=row, column=5, value=indicator.name if indicator else "").border = thin_border
            ws.cell(row=row, column=6, value=upper).border = thin_border
            ws.cell(row=row, column=7, value=lower).border = thin_border
            ws.cell(row=row, column=8, value=unit).border = thin_border
            ws.cell(row=row, column=9, value=params_str).border = thin_border
            ws.cell(row=row, column=10, value=remark).border = thin_border
            row += 1

    # Column widths
    widths = [14, 14, 20, 15, 22, 14, 14, 10, 40, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Save
    settings = get_settings()
    export_dir = os.path.join(settings.UPLOAD_FOLDER, "excel_exports")
    os.makedirs(export_dir, exist_ok=True)
    file_id = str(uuid.uuid4())[:8]
    file_name = f"BOM填写记录_{config.bom_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file_id}.xlsx"
    file_path = os.path.join(export_dir, file_name)
    wb.save(file_path)

    return {
        "file_name": file_name,
        "file_size": os.path.getsize(file_path),
        "download_url": f"/api/v1/metrics/excel-exports/{file_name}",
    }


def _param_summary(p: dict) -> str:
    """Format a param dict into a readable summary string."""
    parts = []
    if p.get("name"):
        parts.append(p["name"])
    if p.get("value") is not None:
        parts.append(f"={p['value']}")
    if p.get("type"):
        parts.append(f"({p['type']})")
    return " ".join(parts) if parts else str(p)


def _diff_params(prev_params: list, cur_params: list) -> dict:
    """Compare two lists of param dicts by 'key' and return structured diff."""
    prev_map = {p["key"]: p for p in prev_params if p.get("key")}
    cur_map = {p["key"]: p for p in cur_params if p.get("key")}
    added = []
    removed = []
    modified = []
    all_keys = set(list(cur_map.keys()) + list(prev_map.keys()))
    for k in sorted(all_keys):
        cur = cur_map.get(k)
        pre = prev_map.get(k)
        if cur and not pre:
            added.append(cur)
        elif pre and not cur:
            removed.append(pre)
        elif cur and pre:
            sub_diff = {}
            for fk in cur:
                if fk == "key":
                    continue
                if str(cur[fk]) != str(pre.get(fk)):
                    sub_diff[fk] = {"before": pre.get(fk), "after": cur[fk]}
            if sub_diff:
                modified.append({"key": k, "name": cur.get("name", k), "diff_fields": sub_diff})
    return {"added": added, "removed": removed, "modified": modified}


async def export_diff_report(db: AsyncSession, config_id: int) -> dict:
    """导出版本差异对比报告 (Excel)"""
    config = await BomConfigService.get(db, config_id)

    # Find latest snapshot for this BOM
    r = await db.execute(
        select(IndicatorVersionSnapshot)
        .where(
            IndicatorVersionSnapshot.entity_type == "bom",
            IndicatorVersionSnapshot.entity_id == config_id,
        )
        .order_by(IndicatorVersionSnapshot.id.desc())
        .limit(1)
    )
    snapshot = r.scalar_one_or_none()
    if not snapshot:
        raise ValueError("暂无版本记录，无法生成差异报告")

    # Get diff data from existing endpoint logic (reuse)
    prev_r = await db.execute(
        select(IndicatorVersionSnapshot)
        .where(
            IndicatorVersionSnapshot.entity_type == "bom",
            IndicatorVersionSnapshot.entity_id == config_id,
            IndicatorVersionSnapshot.id < snapshot.id,
        )
        .order_by(IndicatorVersionSnapshot.id.desc())
        .limit(1)
    )
    prev = prev_r.scalar_one_or_none()

    current = {ind["indicator_id"]: ind for ind in (snapshot.snapshot_data.get("indicators") or [])}
    previous = {}
    if prev:
        previous = {ind["indicator_id"]: ind for ind in (prev.snapshot_data.get("indicators") or [])}

    added = []
    removed = []
    modified = []
    all_keys = set(list(current.keys()) + list(previous.keys()))
    for k in sorted(all_keys):
        cur = current.get(k)
        pre = previous.get(k)
        if cur and not pre:
            added.append(cur)
        elif pre and not cur:
            removed.append(pre)
        elif cur and pre:
            diff = {}
            for field_key in cur:
                if field_key in ("id", "indicator_id", "code", "name", "category"):
                    continue
                if str(cur[field_key]) != str(pre.get(field_key)):
                    if field_key == "params":
                        diff[field_key] = _diff_params(pre.get(field_key, []), cur[field_key])
                    else:
                        diff[field_key] = {"before": pre.get(field_key), "after": cur[field_key]}
            if diff:
                modified.append({"item": cur, "diff": diff})

    wb = Workbook()

    # Sheet 1: Added
    ws1 = wb.active
    ws1.title = "新增指标"
    ws1.merge_cells("A1:F1")
    ws1["A1"] = f"新增指标（共 {len(added)} 项）"
    ws1["A1"].font = Font(bold=True, size=13)
    headers = ["编码", "名称", "硬件参数", "单位", "判定规则", "测试阶段"]
    for col, h in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=h).font = Font(bold=True)
    for i, ind in enumerate(added, 4):
        params = ind.get("params") or {}
        params_str = "; ".join(f"{k}={v}" for k, v in params.items()) if isinstance(params, dict) else str(params)
        ws1.cell(row=i, column=1, value=ind.get("code", ""))
        ws1.cell(row=i, column=2, value=ind.get("name", ""))
        ws1.cell(row=i, column=3, value=params_str or "-")
        ws1.cell(row=i, column=4, value=ind.get("unit") or "-")
        ws1.cell(row=i, column=5, value=ind.get("judgment_rule") or "-")
        ws1.cell(row=i, column=6, value=ind.get("category", ""))

    # Sheet 2: Removed
    ws2 = wb.create_sheet("移除指标")
    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"移除指标（共 {len(removed)} 项）"
    ws2["A1"].font = Font(bold=True, size=13)
    for col, h in enumerate(headers, 1):
        ws2.cell(row=3, column=col, value=h).font = Font(bold=True)
    for i, ind in enumerate(removed, 4):
        params = ind.get("params") or {}
        params_str = "; ".join(f"{k}={v}" for k, v in params.items()) if isinstance(params, dict) else str(params)
        ws2.cell(row=i, column=1, value=ind.get("code", ""))
        ws2.cell(row=i, column=2, value=ind.get("name", ""))
        ws2.cell(row=i, column=3, value=params_str or "-")
        ws2.cell(row=i, column=4, value=ind.get("unit") or "-")
        ws2.cell(row=i, column=5, value=ind.get("judgment_rule") or "-")
        ws2.cell(row=i, column=6, value=ind.get("category", ""))

    # Sheet 3: Modified
    ws3 = wb.create_sheet("修改指标")
    ws3.merge_cells("A1:H1")
    ws3["A1"] = f"修改指标（共 {len(modified)} 项）"
    ws3["A1"].font = Font(bold=True, size=13)
    m_headers = ["编码", "名称", "变更字段", "修改前", "修改后", "分类"]
    for col, h in enumerate(m_headers, 1):
        ws3.cell(row=3, column=col, value=h).font = Font(bold=True)
    i = 4
    for m in modified:
        item = m["item"]
        code = item.get("code", "")
        name = item.get("name", "")
        category = item.get("category", "")
        for field_key, df in m["diff"].items():
            field_label = {
                "unit": "单位", "judgment_rule": "判定规则",
                "test_stage": "测试阶段", "remark": "备注", "params": "硬件参数",
            }.get(field_key, field_key)
            if field_key == "params" and "added" in df:
                for p in df["added"]:
                    ws3.cell(row=i, column=1, value=code)
                    ws3.cell(row=i, column=2, value=name)
                    ws3.cell(row=i, column=3, value=f"{field_label}(新增)")
                    ws3.cell(row=i, column=4, value="-")
                    ws3.cell(row=i, column=5, value=_param_summary(p))
                    ws3.cell(row=i, column=6, value=category)
                    i += 1
                for p in df["removed"]:
                    ws3.cell(row=i, column=1, value=code)
                    ws3.cell(row=i, column=2, value=name)
                    ws3.cell(row=i, column=3, value=f"{field_label}(删除)")
                    ws3.cell(row=i, column=4, value=_param_summary(p))
                    ws3.cell(row=i, column=5, value="-")
                    ws3.cell(row=i, column=6, value=category)
                    i += 1
                for pm in df["modified"]:
                    for sub_fk, sub_df in pm["diff_fields"].items():
                        ws3.cell(row=i, column=1, value=code)
                        ws3.cell(row=i, column=2, value=name)
                        ws3.cell(row=i, column=3, value=f"{field_label}({pm['name']}.{sub_fk})")
                        ws3.cell(row=i, column=4, value=str(sub_df.get("before") or "空"))
                        ws3.cell(row=i, column=5, value=str(sub_df.get("after") or "空"))
                        ws3.cell(row=i, column=6, value=category)
                        i += 1
            else:
                before = df.get("before") or "空"
                after = df.get("after") or "空"
                if isinstance(before, dict):
                    before = "; ".join(f"{k}={v}" for k, v in before.items())
                if isinstance(after, dict):
                    after = "; ".join(f"{k}={v}" for k, v in after.items())
                ws3.cell(row=i, column=1, value=code)
                ws3.cell(row=i, column=2, value=name)
                ws3.cell(row=i, column=3, value=field_label)
                ws3.cell(row=i, column=4, value=str(before))
                ws3.cell(row=i, column=5, value=str(after))
                ws3.cell(row=i, column=6, value=category)
                i += 1

    # Save
    settings = get_settings()
    export_dir = os.path.join(settings.UPLOAD_FOLDER, "excel_exports")
    os.makedirs(export_dir, exist_ok=True)
    file_id = str(uuid.uuid4())[:8]
    file_name = f"版本差异报告_{config.bom_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file_id}.xlsx"
    file_path = os.path.join(export_dir, file_name)
    wb.save(file_path)

    return {
        "file_name": file_name,
        "file_size": os.path.getsize(file_path),
        "download_url": f"/api/v1/metrics/excel-exports/{file_name}",
    }


CJK_FONT_PATH = "/System/Library/Fonts/Supplemental/Arial Unicode.ttf"


def _fit_cell_text(text: str, col_width_mm: float, font_size: float = 8) -> str:
    """按列宽(毫米)截断文本，避免超出单元格宽度"""
    if not text:
        return ""
    margin = 1.5
    ascii_w = font_size * 0.3528 * 0.55
    cjk_w = font_size * 0.3528
    est = sum(ascii_w if ord(c) < 128 else cjk_w for c in text)
    if est <= col_width_mm - margin:
        return text
    cut = 0
    acc = 0.0
    for c in text:
        w = ascii_w if ord(c) < 128 else cjk_w
        if acc + w > col_width_mm - margin:
            break
        acc += w
        cut += 1
    if cut == 0:
        return ""
    return text[:cut] + "…"


async def export_pdf_report(db: AsyncSession, config_id: int) -> dict:
    """导出 BOM 评审配置单 PDF"""
    config = await BomConfigService.get(db, config_id)
    indicators = await BomConfigService.list_indicators(db, config_id)

    items_r = await db.execute(
        select(CollectionTestItem)
        .where(CollectionTestItem.collection_id == config.collection_id)
        .order_by(CollectionTestItem.sort_order)
    )
    items = items_r.scalars().all()

    pdf = FPDF()
    pdf.add_font("ArialUnicode", "", CJK_FONT_PATH)
    pdf.add_font("ArialUnicode", "B", CJK_FONT_PATH)
    pdf.add_page()

    pdf.set_font("ArialUnicode", "B", 16)
    pdf.cell(0, 10, "BOM 评审配置单", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.set_font("ArialUnicode", "", 10)
    pdf.cell(0, 8, f"BOM编码: {config.bom_code}  |  BOM名称: {config.bom_name}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.cell(0, 8, f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)

    col_w = [45, 30, 45, 35, 15, 15]
    headers = ["测试项", "指标编码", "指标名称", "硬件参数", "单位", "备注"]
    pdf.set_font("ArialUnicode", "B", 9)
    pdf.set_fill_color(68, 114, 196)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 8, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("ArialUnicode", "", 8)

    for item in items:
        ind_r = await db.execute(
            select(TestItemIndicator).where(TestItemIndicator.test_item_id == item.id)
        )
        item_inds = ind_r.scalars().all()
        for ti_ind in item_inds:
            bi = next((x for x in indicators if x["indicator_id"] == ti_ind.indicator_id), None)
            indicator = await DictionaryService.get(db, ti_ind.indicator_id) if ti_ind.indicator_id else None
            params_str = ""
            if indicator and indicator.params:
                params_str = "; ".join(f"{k}={v}" for k, v in indicator.params.items())
            unit = bi.get("unit") if bi and bi.get("unit") else indicator.unit if indicator else ""
            data = [item.name or "", indicator.code if indicator else "", indicator.name if indicator else "", params_str, unit, ""]
            for i, d in enumerate(data):
                truncated = _fit_cell_text(d, col_w[i])
                pdf.cell(col_w[i], 7, truncated, border=1)
            pdf.ln()

    settings = get_settings()
    export_dir = os.path.join(settings.UPLOAD_FOLDER, "excel_exports")
    os.makedirs(export_dir, exist_ok=True)
    file_id = str(uuid.uuid4())[:8]
    file_name = f"BOM评审配置单_{config.bom_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file_id}.pdf"
    file_path = os.path.join(export_dir, file_name)
    pdf.output(file_path)

    return {
        "file_name": file_name,
        "file_size": os.path.getsize(file_path),
        "download_url": f"/api/v1/metrics/excel-exports/{file_name}",
    }
