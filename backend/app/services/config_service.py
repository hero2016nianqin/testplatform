"""
ConfigManager — 配置导入导出服务
支持 CSV / XLSX / JSON / XML 四种格式
对应 design.md §5.1, §9
"""
import csv
import io
import json
import xml.etree.ElementTree as ET
from typing import List, Dict, Any, Optional

import pandas as pd
from fastapi import UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.test_item import TestItem
from app.core.exceptions import BusinessException


class ConfigImportError(BusinessException):
    def __init__(self, message: str):
        super().__init__(code=400, message=message)


class ConfigService:
    SUPPORTED_FORMATS = {"csv", "xlsx", "json", "xml"}

    @staticmethod
    async def parse_import(file: UploadFile, fmt: str) -> Dict[str, Any]:
        if fmt not in ConfigService.SUPPORTED_FORMATS:
            raise ConfigImportError(f"不支持的格式: {fmt}，支持: {', '.join(ConfigService.SUPPORTED_FORMATS)}")
        content = await file.read()
        if fmt == "csv":
            return ConfigService._parse_csv(content)
        elif fmt == "xlsx":
            return ConfigService._parse_xlsx(content)
        elif fmt == "json":
            return ConfigService._parse_json(content)
        elif fmt == "xml":
            return ConfigService._parse_xml(content)

    @staticmethod
    def _parse_csv(content: bytes) -> Dict[str, Any]:
        stream = io.StringIO(content.decode("utf-8"))
        reader = csv.DictReader(stream)
        items = list(reader)
        return {"format": "csv", "items": items, "columns": reader.fieldnames, "count": len(items)}

    @staticmethod
    def _parse_xlsx(content: bytes) -> Dict[str, Any]:
        df = pd.read_excel(io.BytesIO(content))
        items = df.to_dict(orient="records")
        return {"format": "xlsx", "items": items, "columns": list(df.columns), "count": len(items)}

    @staticmethod
    def _parse_json(content: bytes) -> Dict[str, Any]:
        data = json.loads(content)
        items = data if isinstance(data, list) else data.get("items", [])
        return {"format": "json", "items": items, "count": len(items)}

    @staticmethod
    def _parse_xml(content: bytes) -> Dict[str, Any]:
        root = ET.fromstring(content)
        items = [{elem.tag: elem.text for elem in child} for child in root]
        return {"format": "xml", "items": items, "count": len(items)}

    @staticmethod
    def validate_items(items: List[Dict]) -> Dict[str, Any]:
        required = {"name", "expected_value", "min_value", "max_value"}
        validated = []
        errors = []
        for idx, item in enumerate(items):
            missing = required - set(item.keys())
            if missing:
                errors.append(f"第{idx + 1}行: 缺少字段 {missing}")
                continue
            try:
                validated.append({
                    "name": str(item["name"]).strip(),
                    "description": str(item.get("description", "")).strip(),
                    "expected_value": float(item["expected_value"]),
                    "min_value": float(item["min_value"]),
                    "max_value": float(item["max_value"]),
                    "unit": str(item.get("unit", "")).strip(),
                    "category": str(item.get("category", "general")).strip(),
                })
            except (ValueError, TypeError) as e:
                errors.append(f"第{idx + 1}行: 数值格式错误 - {e}")
        return {"validated": validated, "errors": errors, "total": len(items), "valid_count": len(validated), "error_count": len(errors)}

    @staticmethod
    async def import_items(db: AsyncSession, items: List[Dict]) -> int:
        count = 0
        for data in items:
            existing = await db.execute(select(TestItem).where(TestItem.name == data["name"]))
            if existing.scalar_one_or_none():
                continue
            db.add(TestItem(**data))
            count += 1
        await db.flush()
        return count

    @staticmethod
    async def export_items(db: AsyncSession, fmt: str) -> Any:
        result = await db.execute(select(TestItem).order_by(TestItem.sort_order))
        items = list(result.scalars().all())
        if fmt == "json":
            data = {"export_time": __import__("datetime").datetime.now().isoformat(), "items": [i.to_dict() for i in items]}
            return json.dumps(data, ensure_ascii=False, indent=2), "application/json"
        elif fmt == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(["name", "description", "expected_value", "min_value", "max_value", "unit", "category"])
            for i in items:
                writer.writerow([i.name, i.description, i.expected_value, i.min_value, i.max_value, i.unit, i.category])
            return output.getvalue(), "text/csv"
        elif fmt == "xlsx":
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ["name", "description", "expected_value", "min_value", "max_value", "unit", "category"]
            bold = Font(bold=True)
            thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = bold
                c.border = thin
                c.alignment = Alignment(horizontal="center")
            for ri, item in enumerate(items, 2):
                for ci, field in enumerate(headers, 1):
                    c = ws.cell(row=ri, column=ci, value=getattr(item, field, ""))
                    c.border = thin
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            raise ConfigImportError(f"不支持的导出格式: {fmt}")
